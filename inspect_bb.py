import openpyxl

path = "data/raw/bangladesh_bank/2026/08/23/time_series_data1972-2024.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

print(f"Number of sheets: {len(wb.sheetnames)}\n")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  - {name!r}  (rows: {ws.max_row}, cols: {ws.max_column})")
